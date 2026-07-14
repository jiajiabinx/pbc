"""Tool schemas + implementations for the agent loop.

Every capability — parsing, OCR, matching, extraction, verification, status
updates, escalation — is a tool the model chooses to call. Guardrails live in
the implementations, not in prompts:
  * update_item_status is rejected without a verify_item verdict on record
  * extract_fields validates every citation against the actual source document
  * verify_item re-derives sufficiency from the raw document — it never sees the
    agent's claims
"""
from __future__ import annotations

import json
import base64
import zipfile
from dataclasses import dataclass, field
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, coordinate_to_tuple

import models
import versioning
from store import Store, StatusGuardError

MAX_DOC_CHARS = 7000


# --------------------------------------------------------------------- schemas

TOOL_SCHEMAS = [
    {
        "name": "submit_plan",
        "description": "REQUIRED first call of every episode. Classify the email and state your plan before doing anything else.",
        "input_schema": {
            "type": "object",
            "properties": {
                "classification": {
                    "type": "string",
                    "enum": ["client_documents", "auditor_request", "acknowledgement_or_no_action",
                             "clarification_needed", "other"],
                    "description": "What kind of email this is.",
                },
                "steps": {"type": "array", "items": {"type": "string"},
                          "description": "The ordered steps you intend to take (can be ['do nothing'])."},
                "items_possibly_affected": {"type": "array", "items": {"type": "string"},
                                            "description": "PBC item ids this email might affect, if any."},
            },
            "required": ["classification", "steps"],
        },
    },
    {
        "name": "register_document",
        "description": "Register an email attachment into the tracker before parsing it. Handles version lineage (e.g. Final_v3_REAL.xlsx superseding v2). Returns doc_id.",
        "input_schema": {
            "type": "object",
            "properties": {
                "attachment_ref": {"type": "string", "description": "The att_N reference from the email context."},
            },
            "required": ["attachment_ref"],
        },
    },
    {
        "name": "parse_pdf",
        "description": "Extract text from a registered PDF document, with [page N] markers.",
        "input_schema": {
            "type": "object",
            "properties": {"doc_id": {"type": "integer"}},
            "required": ["doc_id"],
        },
    },
    {
        "name": "parse_excel",
        "description": "Read a registered Excel document: sheet names, dimensions, and cell values with A1-style references.",
        "input_schema": {
            "type": "object",
            "properties": {"doc_id": {"type": "integer"},
                           "sheet": {"type": "string", "description": "Optional sheet name; default first sheet."}},
            "required": ["doc_id"],
        },
    },
    {
        "name": "ocr_image",
        "description": "Transcribe an image/photo or scanned PDF (e.g. a whiteboard photo, a scan with no text layer) using a vision model. Costs more than parsing — prefer parse_pdf/parse_excel for digital documents.",
        "input_schema": {
            "type": "object",
            "properties": {"doc_id": {"type": "integer"}},
            "required": ["doc_id"],
        },
    },
    {
        "name": "unzip",
        "description": "Extract a registered zip archive; each contained file is registered as its own document. Returns the new doc_ids.",
        "input_schema": {
            "type": "object",
            "properties": {"doc_id": {"type": "integer"}},
            "required": ["doc_id"],
        },
    },
    {
        "name": "match_pbc_items",
        "description": "Find candidate PBC items for a document/description using local embeddings. Use this instead of guessing which item a document answers.",
        "input_schema": {
            "type": "object",
            "properties": {"query": {"type": "string",
                                     "description": "Filename + a short description of the document content."},
                           "top_k": {"type": "integer", "default": 5}},
            "required": ["query"],
        },
    },
    {
        "name": "extract_fields",
        "description": "Extract key fields (period, entity, dates, totals) from a registered document with citations. Each citation is validated against the source; invalid citations are rejected.",
        "input_schema": {
            "type": "object",
            "properties": {
                "doc_id": {"type": "integer"},
                "fields": {"type": "array", "items": {"type": "string"},
                           "description": "Field names to extract, e.g. ['period_end', 'entity', 'total_amount']."},
            },
            "required": ["doc_id", "fields"],
        },
    },
    {
        "name": "verify_item",
        "description": "Run the independent verifier: does this document actually satisfy the item's acceptance criteria? REQUIRED before setting status to Received/Insufficient/Complete. The verifier reads the raw document itself.",
        "input_schema": {
            "type": "object",
            "properties": {
                "item_id": {"type": "string"},
                "doc_id": {"type": "integer"},
                "context": {"type": "string",
                            "description": "Optional thread context the verifier should know, e.g. 'auditor said 7 of 10 confirmations still outstanding'. State facts from emails only, not your own conclusions."},
            },
            "required": ["item_id", "doc_id"],
        },
    },
    {
        "name": "update_item_status",
        "description": "Update a PBC item's tracker status. Received/Insufficient/Complete require a prior verify_item verdict for the item (enforced in code).",
        "input_schema": {
            "type": "object",
            "properties": {
                "item_id": {"type": "string"},
                "status": {"type": "string",
                           "enum": ["Not started", "Requested", "Under review", "Received",
                                    "Insufficient", "Complete"]},
                "confidence": {"type": "number", "description": "0-1"},
                "rationale": {"type": "string"},
                "doc_id": {"type": "integer", "description": "The document supporting this status, if any."},
            },
            "required": ["item_id", "status", "rationale"],
        },
    },
    {
        "name": "flag_clarification",
        "description": "Record that a clarification question should go back to the sender (feeds the follow-up drafts).",
        "input_schema": {
            "type": "object",
            "properties": {
                "item_id": {"type": "string", "description": "Related PBC item, if any."},
                "question": {"type": "string"},
                "recipient": {"type": "string", "description": "Email address of who should answer."},
            },
            "required": ["question", "recipient"],
        },
    },
    {
        "name": "escalate",
        "description": "Hand this email to a more capable (more expensive) model because it is ambiguous or high-stakes. State why. Use sparingly.",
        "input_schema": {
            "type": "object",
            "properties": {"reason": {"type": "string"}},
            "required": ["reason"],
        },
    },
    {
        "name": "done",
        "description": "Finish this episode with a one-line summary of what changed.",
        "input_schema": {
            "type": "object",
            "properties": {"summary": {"type": "string"}},
            "required": ["summary"],
        },
    },
]


# --------------------------------------------------------------------- context

@dataclass
class ToolContext:
    store: Store
    router: models.Router
    matcher: object            # embeddings.ItemMatcher
    profile_text: str
    email: object              # ingest.Email
    episode_id: int
    attachments: dict = field(default_factory=dict)   # att_N -> ingest.Attachment
    doc_cache: dict = field(default_factory=dict)     # doc_id -> parsed text


class Escalate(Exception):
    def __init__(self, reason: str):
        self.reason = reason


class Done(Exception):
    def __init__(self, summary: str):
        self.summary = summary


# --------------------------------------------------------------- doc rendering

def _render_pdf(path: str) -> str:
    out = []
    with pdfplumber.open(path) as pdf:
        for i, page in enumerate(pdf.pages, 1):
            out.append(f"[page {i}]\n{(page.extract_text() or '').strip()}")
    return "\n\n".join(out)


def _render_excel(path: str, sheet: str | None = None, max_rows: int = 80, max_cols: int = 16) -> str:
    wb = load_workbook(path, data_only=True, read_only=True)
    out = [f"Sheets: {', '.join(wb.sheetnames)}"]
    names = [sheet] if sheet and sheet in wb.sheetnames else wb.sheetnames
    for name in names:
        ws = wb[name]
        out.append(f"\n[sheet {name}] dimensions={ws.max_row}x{ws.max_column}")
        for r, row in enumerate(ws.iter_rows(max_row=min(ws.max_row or 1, max_rows),
                                             max_col=min(ws.max_column or 1, max_cols)), 1):
            cells = [f"{c.coordinate}={c.value}" for c in row if c.value is not None]
            if cells:
                out.append("  " + " | ".join(cells))
    wb.close()
    return "\n".join(out)


def _doc_content(ctx: ToolContext, doc_id: int) -> tuple[str, str]:
    """Return (kind, text) for a registered doc, cached per episode run."""
    if doc_id in ctx.doc_cache:
        return ctx.doc_cache[doc_id]
    doc = ctx.store.get_document(doc_id)
    if doc is None:
        raise ValueError(f"Unknown doc_id {doc_id}")
    fam = versioning.ext_family(doc["filename"])
    if fam == "pdf":
        text = _render_pdf(doc["path"])
        # A scanned PDF has page markers but (almost) no extractable text —
        # fall back to vision transcription so 'quality varies' scans still work.
        pages = max(text.count("[page "), 1)
        if len(text) - pages * len("[page N]") < 40 * pages:
            result = ("pdf", _ocr(ctx, doc))
        else:
            result = ("pdf", text)
    elif fam == "excel":
        result = ("excel", _render_excel(doc["path"]))
    elif fam == "image":
        result = ("image", _ocr(ctx, doc))
    elif fam == "text":
        result = ("text", Path(doc["path"]).read_text(errors="replace"))
    else:
        raise ValueError(f"Cannot render document family {fam!r} ({doc['filename']}); "
                         "if it is a zip, call unzip first.")
    ctx.doc_cache[doc_id] = result
    return result


_MEDIA = {".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png",
          ".gif": "image/gif", ".webp": "image/webp"}


def _ocr(ctx: ToolContext, doc) -> str:
    """Vision transcription (images and scanned PDFs), cached in the DB by
    content hash so we never pay twice."""
    cached = ctx.store.get_meta(f"ocr:{doc['sha256']}")
    if cached:
        return cached
    data = base64.standard_b64encode(Path(doc["path"]).read_bytes()).decode()
    if versioning.ext_family(doc["filename"]) == "pdf":
        content_block = {"type": "document",
                         "source": {"type": "base64", "media_type": "application/pdf",
                                    "data": data}}
    else:
        media = _MEDIA.get(Path(doc["path"]).suffix.lower(), "image/jpeg")
        content_block = {"type": "image",
                         "source": {"type": "base64", "media_type": media, "data": data}}
    resp = ctx.router.call(
        models.VISION, purpose=f"ocr:{doc['filename']}", episode_id=ctx.episode_id,
        max_tokens=1500,
        messages=[{"role": "user", "content": [
            content_block,
            {"type": "text", "text":
                "Transcribe everything in this document image faithfully: all text, numbers, "
                "tables, dates, account names, signatures, tick marks. Note the document's "
                "physical nature (typed report / handwritten / whiteboard photo / scan). "
                "Do not interpret or summarize — transcribe."},
        ]}],
    )
    text = "".join(b.text for b in resp.content if b.type == "text")
    ctx.store.set_meta(f"ocr:{doc['sha256']}", text)
    return text


# ------------------------------------------------------------- implementations

def _tool_submit_plan(ctx: ToolContext, inp: dict) -> str:
    ctx.store.add_trace(ctx.episode_id, "plan", "submit_plan", inp)
    return "Plan recorded. Proceed."


def _tool_register_document(ctx: ToolContext, inp: dict) -> str:
    att = ctx.attachments.get(inp["attachment_ref"])
    if att is None:
        return json.dumps({"error": f"No attachment {inp['attachment_ref']!r} on this email. "
                                    f"Available: {sorted(ctx.attachments)}"})
    key = versioning.semantic_key(att.filename)
    info = ctx.store.register_document(att.filename, att.path, att.sha256,
                                       ctx.email.email_id, key)
    info["filename"] = att.filename
    info["semantic_key"] = key
    if info["supersedes"]:
        prev = ctx.store.get_document(info["supersedes"])
        info["supersedes_filename"] = prev["filename"] if prev else None
        info["note"] = "This file supersedes an earlier version of the same document."
    if info["duplicate"]:
        info["note"] = "Exact duplicate of an already-registered document (same content hash)."
    return json.dumps(info)


def _tool_parse_pdf(ctx: ToolContext, inp: dict) -> str:
    kind, text = _doc_content(ctx, int(inp["doc_id"]))
    if kind == "image":
        return "This is an image, transcribed via OCR:\n" + text[:MAX_DOC_CHARS]
    return text[:MAX_DOC_CHARS]


def _tool_parse_excel(ctx: ToolContext, inp: dict) -> str:
    doc = ctx.store.get_document(int(inp["doc_id"]))
    if doc is None:
        return json.dumps({"error": f"Unknown doc_id {inp['doc_id']}"})
    return _render_excel(doc["path"], inp.get("sheet"))[:MAX_DOC_CHARS]


def _tool_ocr_image(ctx: ToolContext, inp: dict) -> str:
    doc = ctx.store.get_document(int(inp["doc_id"]))
    if doc is None:
        return json.dumps({"error": f"Unknown doc_id {inp['doc_id']}"})
    return _ocr(ctx, doc)[:MAX_DOC_CHARS]


def _tool_unzip(ctx: ToolContext, inp: dict) -> str:
    doc = ctx.store.get_document(int(inp["doc_id"]))
    if doc is None:
        return json.dumps({"error": f"Unknown doc_id {inp['doc_id']}"})
    out_dir = Path(doc["path"]).parent / (Path(doc["path"]).stem + "_extracted")
    out_dir.mkdir(exist_ok=True)
    children = []
    with zipfile.ZipFile(doc["path"]) as zf:
        for name in zf.namelist():
            if name.endswith("/") or name.startswith("__MACOSX"):
                continue
            safe = Path(name).name
            data = zf.read(name)
            p = out_dir / safe
            p.write_bytes(data)
            import hashlib
            info = ctx.store.register_document(
                safe, str(p), hashlib.sha256(data).hexdigest(),
                ctx.email.email_id, versioning.semantic_key(safe),
                parent_doc_id=doc["doc_id"])
            children.append({"doc_id": info["doc_id"], "filename": safe,
                             "version": info["version"]})
    return json.dumps({"extracted": children, "count": len(children)})


def _tool_match_pbc_items(ctx: ToolContext, inp: dict) -> str:
    return json.dumps(ctx.matcher.match(inp["query"], int(inp.get("top_k", 5))))


_EXTRACT_SCHEMA = {
    "type": "object",
    "properties": {
        "fields": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "name": {"type": "string"},
                    "value": {"type": "string"},
                    "citation": {"type": "string",
                                 "description": "Exactly 'page N' for PDFs/images or a cell "
                                                "reference like 'SheetName!B4' for Excel."},
                    "found": {"type": "boolean"},
                },
                "required": ["name", "value", "citation", "found"],
                "additionalProperties": False,
            },
        },
    },
    "required": ["fields"],
    "additionalProperties": False,
}


def _validate_citation(kind: str, citation: str, doc_path: str, text: str) -> bool:
    citation = citation.strip()
    if kind in ("pdf", "image"):
        if not citation.lower().startswith("page"):
            return False
        try:
            n = int(citation.split()[1])
        except (IndexError, ValueError):
            return False
        pages = text.count("[page ") or 1
        return 1 <= n <= pages
    if kind == "excel":
        ref = citation.split("!")[-1].replace("$", "")
        try:
            row, col = coordinate_to_tuple(ref)
        except Exception:
            return False
        try:
            wb = load_workbook(doc_path, read_only=True)
            sheet = citation.split("!")[0] if "!" in citation else wb.sheetnames[0]
            if sheet not in wb.sheetnames:
                return False
            ws = wb[sheet]
            ok = row <= (ws.max_row or 0) and col <= (ws.max_column or 0)
            wb.close()
            return ok
        except Exception:
            return False
    return True  # text docs: no positional citation to validate


def _tool_extract_fields(ctx: ToolContext, inp: dict) -> str:
    doc_id = int(inp["doc_id"])
    doc = ctx.store.get_document(doc_id)
    if doc is None:
        return json.dumps({"error": f"Unknown doc_id {doc_id}"})
    kind, text = _doc_content(ctx, doc_id)
    payload = models.structured_json(
        ctx.router, models.EXTRACTOR, purpose=f"extract:{doc['filename']}",
        episode_id=ctx.episode_id, max_tokens=1500, schema=_EXTRACT_SCHEMA,
        user=(f"Document: {doc['filename']}\n\n{text[:MAX_DOC_CHARS]}\n\n"
              f"Extract these fields: {inp['fields']}. For each, give the exact value as it "
              f"appears and a citation ('page N' or 'Sheet!Cell'). If a field is absent, set "
              f"found=false and value=''."))
    for f in payload["fields"]:
        if f["found"] and not _validate_citation(kind, f["citation"], doc["path"], text):
            f["found"] = False
            f["rejected"] = "citation does not exist in source document"
    ctx.store.add_trace(ctx.episode_id, "tool_result", "extract_fields_validated", payload)
    return json.dumps(payload)


_VERIFY_SCHEMA = {
    "type": "object",
    "properties": {
        "verdict": {"type": "string", "enum": ["sufficient", "insufficient"]},
        "criteria": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "criterion": {"type": "string"},
                    "met": {"type": "boolean"},
                    "evidence": {"type": "string"},
                },
                "required": ["criterion", "met", "evidence"],
                "additionalProperties": False,
            },
        },
        "rationale": {"type": "string"},
        "confidence": {"type": "number"},
    },
    "required": ["verdict", "criteria", "rationale", "confidence"],
    "additionalProperties": False,
}


def _tool_verify_item(ctx: ToolContext, inp: dict) -> str:
    item = ctx.store.get_item(inp["item_id"])
    doc = ctx.store.get_document(int(inp["doc_id"]))
    if item is None or doc is None:
        return json.dumps({"error": "unknown item_id or doc_id"})
    kind, text = _doc_content(ctx, int(inp["doc_id"]))
    thread_context = inp.get("context", "")
    system = (
        "You are an independent audit evidence verifier at a CPA firm. You are given a PBC "
        "line item with acceptance criteria, the client profile that defines what 'correct' "
        "means (entities, fiscal year), and the raw content of one submitted document. "
        "Judge strictly whether THIS document satisfies the acceptance criteria. Common "
        "failure modes to check: wrong period (e.g. Q2 sent when Q3 asked), wrong or "
        "missing entity (consolidated means ALL listed entities), partial coverage (3 of "
        "10 confirmations, one bank account of several), informal evidence (handwritten/"
        "whiteboard where a formal document is required), unsigned where signature is "
        "required. Partial coverage of a multi-part item is insufficient. Evaluate each "
        "acceptance criterion separately with evidence citations."
    )
    user = (
        f"CLIENT PROFILE:\n{ctx.profile_text}\n\n"
        f"PBC ITEM {item['item_id']} [{item['category']}] priority={item['priority']}\n"
        f"Request: {item['description']}\n"
        f"Acceptance criteria: {item['acceptance']}\n"
        f"Expected document types: {item['expected_docs']}\n\n"
        f"DOCUMENT: {doc['filename']} ({kind})\n{text[:MAX_DOC_CHARS]}\n\n"
        + (f"THREAD FACTS (from the email thread, verify against them): {thread_context}\n\n"
           if thread_context else "")
        + "Does this document satisfy the acceptance criteria?"
    )
    verdict = models.structured_json(
        ctx.router, models.VERIFIER, purpose=f"verify:{item['item_id']}",
        episode_id=ctx.episode_id, max_tokens=2500,
        schema=_VERIFY_SCHEMA, system=system, user=user)
    ctx.store.add_verification(item["item_id"], doc["doc_id"], verdict["verdict"],
                               verdict["rationale"], verdict["criteria"],
                               verdict["confidence"], ctx.episode_id)
    ctx.store.add_trace(ctx.episode_id, "verdict", f"verify:{item['item_id']}", verdict)
    return json.dumps(verdict)


def _tool_update_item_status(ctx: ToolContext, inp: dict) -> str:
    try:
        ctx.store.update_item_status(
            inp["item_id"], inp["status"],
            confidence=inp.get("confidence"), rationale=inp.get("rationale"),
            doc_id=inp.get("doc_id"), email_id=ctx.email.email_id)
    except (StatusGuardError, ValueError) as e:
        return json.dumps({"error": str(e)})
    return json.dumps({"ok": True, "item_id": inp["item_id"], "status": inp["status"]})


def _tool_flag_clarification(ctx: ToolContext, inp: dict) -> str:
    ctx.store.add_clarification(inp.get("item_id"), inp["question"], inp["recipient"],
                                ctx.email.email_id, ctx.episode_id)
    return json.dumps({"ok": True})


def _tool_escalate(ctx: ToolContext, inp: dict) -> str:
    raise Escalate(inp["reason"])


def _tool_done(ctx: ToolContext, inp: dict) -> str:
    raise Done(inp.get("summary", ""))


_IMPLS = {
    "submit_plan": _tool_submit_plan,
    "register_document": _tool_register_document,
    "parse_pdf": _tool_parse_pdf,
    "parse_excel": _tool_parse_excel,
    "ocr_image": _tool_ocr_image,
    "unzip": _tool_unzip,
    "match_pbc_items": _tool_match_pbc_items,
    "extract_fields": _tool_extract_fields,
    "verify_item": _tool_verify_item,
    "update_item_status": _tool_update_item_status,
    "flag_clarification": _tool_flag_clarification,
    "escalate": _tool_escalate,
    "done": _tool_done,
}


def dispatch(ctx: ToolContext, name: str, inp: dict) -> str:
    """Execute one tool call; Escalate/Done propagate to the loop."""
    impl = _IMPLS.get(name)
    if impl is None:
        return json.dumps({"error": f"unknown tool {name!r}"})
    try:
        return impl(ctx, inp)
    except (Escalate, Done):
        raise
    except Exception as e:  # tool errors go back to the model, not up the stack
        return json.dumps({"error": f"{type(e).__name__}: {e}"})
